#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/4/25 21:40
# @Author  : Mr.chen
# @Site    : 
# @File    : do_excel.py
# @Software: PyCharm
# @Email   : 794281961@qq.com


# 把测试数据从Excel读取出来
from openpyxl import load_workbook
from selenium import webdriver


class DoExcel:
    actually = 9
    result = 10

    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_data(self):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        # 把每行的测试数据都放在一个list里面、然后所有行的数据存在一个大list里面
        test_data = []  # 存储所有行的数据
        for i in range(2, sheet.max_row + 1):
            sub_data = []
            for j in range(1, 9):
                sub_data.append(sheet.cell(i, j).value)
                # print(sub_data)
            test_data.append(sub_data)
        return test_data

    def write_data(self, row, actual, result):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        sheet.cell(row, self.actually).value = actual
        sheet.cell(row, self.result).value = result
        wb.save(self.file_path)


if __name__ == '__main__':
    from conf.read_config import ReadConfig

    data_path = ReadConfig().read_path('/project.conf', 'PROJECTS', 'test_data_path')
    print(DoExcel(data_path, 'test_data').read_data())
